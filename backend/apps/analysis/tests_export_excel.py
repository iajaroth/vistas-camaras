"""Tests for Excel export endpoints and functions."""

from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.analysis.exporters import export_camera_excel, export_group_excel
from apps.analysis.models import AnalysisReport, CriticalNote
from apps.cameras.models import Camera, CameraGroup

User = get_user_model()


class ExcelExporterTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="op", password="pass")
        self.group = CameraGroup.objects.create(
            code="EX01", name="Excel Test", next_sequence=3
        )
        self.cam1 = Camera.objects.create(
            group=self.group,
            compound_code="EX01-01",
            name="Cam 1",
            day_view_path="EX01/EX01-01_day.jpg",
        )
        self.cam2 = Camera.objects.create(
            group=self.group,
            compound_code="EX01-02",
            name="Cam 2",
            night_view_path="EX01/EX01-02_night.jpg",
        )
        AnalysisReport.objects.create(
            camera=self.cam1,
            pros=["Buena cobertura"],
            improvements=["Ajustar ángulo"],
            recommended_analytics=["Detección de movimiento"],
            critical_notes=["Zona ciega lateral"],
        )
        CriticalNote.objects.create(
            camera=self.cam2, author=self.user, content="Cable expuesto"
        )

    def test_export_group_excel_content(self):
        buf = export_group_excel(self.group)
        wb = load_workbook(BytesIO(buf.getvalue()))
        ws = wb.active

        self.assertEqual(ws.title, "Grupo EX01")
        # Header row + 2 data rows
        self.assertEqual(ws.max_row, 3)
        self.assertEqual(ws.max_column, 9)

        # First camera row
        row1 = [cell.value for cell in ws[2]]
        self.assertEqual(row1[0], "EX01-01")
        self.assertEqual(row1[2], "Sí")  # day view
        self.assertEqual(row1[3], "No")  # night view
        self.assertIn("Buena cobertura", row1[4])
        self.assertEqual(row1[8], "Analizada")

        # Second camera row
        row2 = [cell.value for cell in ws[3]]
        self.assertEqual(row2[0], "EX01-02")
        self.assertEqual(row2[2], "No")  # no day view
        self.assertEqual(row2[3], "Sí")  # night view
        self.assertIn("Cable expuesto", row2[7])
        self.assertEqual(row2[8], "Pendiente")

    def test_export_camera_excel_single(self):
        buf = export_camera_excel(self.cam1)
        wb = load_workbook(BytesIO(buf.getvalue()))
        ws = wb.active

        self.assertEqual(ws.title, "EX01-01")
        self.assertEqual(ws.max_row, 2)  # header + 1 data row
        row = [cell.value for cell in ws[2]]
        self.assertEqual(row[0], "EX01-01")
        self.assertEqual(row[8], "Analizada")


class ExcelEndpointTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="op", password="pass")
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        self.group = CameraGroup.objects.create(
            code="EP01", name="Endpoint Test", next_sequence=2
        )
        self.cam = Camera.objects.create(
            group=self.group, compound_code="EP01-01", name="Cam EP"
        )

    def test_group_export_excel_endpoint(self):
        resp = self.client.get(f"/api/groups/{self.group.pk}/export/excel/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("EP01.xlsx", resp["Content-Disposition"])

    def test_camera_export_excel_endpoint(self):
        resp = self.client.get(f"/api/cameras/{self.cam.pk}/export/excel/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("EP01-01.xlsx", resp["Content-Disposition"])

    def test_export_excel_requires_auth(self):
        client = APIClient()
        resp = client.get(f"/api/groups/{self.group.pk}/export/excel/")
        self.assertEqual(resp.status_code, 401)
